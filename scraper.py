#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Hygiene.bg -> TEMU general upload template scraper

Collects products from hygiene.bg and fills the uploaded TEMU_GENERAL_UPLOAD_TEMPLATE.xlsx
Template sheet. Variable WooCommerce products are exported with one row per variation.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import logging
import re
import shutil
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

BASE_URL = "https://hygiene.bg/"
USER_AGENT = "Mozilla/5.0 (compatible; HygieneTemuScraper/1.0; +https://hygiene.bg/)"

URL_RE = re.compile(r"^https?://", re.I)
PRICE_RE = re.compile(r"([0-9]+(?:[\s,.][0-9]{1,2})?)\s*(?:€|&euro;|eur\b)", re.I)
SIZE_RE = re.compile(
    r"(?<![\w])(?P<num>\d+(?:[\.,]\d+)?)\s*(?P<unit>мл\.?|ml\.?|мил[иі]лит(?:ъра?|ра)?|л\.?|l\.?|лит(?:ър|ра|ри)?|гр\.?|г\.?|g\.?|kg\.?|кг\.?)\b",
    re.I,
)
THUMB_RE = re.compile(r"-(?:64|100|150|300|320|450|600|768|800)x(?:64|100|150|300|320|450|600|768|800)(?=\.)")

SKIP_IMAGE_PARTS = (
    "/logo",
    "logo-",
    "favico",
    "top-seller",
    "top-quality",
    "ajax-loader",
    "32x32.gif",
    "placeholder",
)


@dataclass
class ProductRow:
    product_url: str = ""
    listing_id: str = ""
    seller_sku: str = ""
    item_name: str = ""
    category: str = ""
    brand: str = ""
    trademark: str = ""
    variation_type_1: str = ""
    variation_variant_1: str = ""
    variation_type_2: str = ""
    variation_variant_2: str = ""
    item_notes: list[str] = field(default_factory=list)
    item_description: str = ""
    keyword_attributes: str = ""
    images: list[str] = field(default_factory=list)
    map_price: Optional[float] = None
    net_content: str = ""
    total_net_content: str = ""
    net_content_unit: str = ""
    availability: str = ""
    raw_variation: str = ""


def clean_text(value: Any, max_len: Optional[int] = None) -> str:
    if value is None:
        return ""
    text = html.unescape(str(value))
    text = re.sub(r"\s+", " ", text).strip()
    if max_len and len(text) > max_len:
        return text[: max_len - 1].rstrip() + "…"
    return text


def parse_price_text(text: str) -> Optional[float]:
    if not text:
        return None
    text = html.unescape(text).replace("\xa0", " ")
    m = PRICE_RE.search(text)
    if not m:
        return None
    raw = m.group(1).replace(" ", "").replace(",", ".")
    try:
        return round(float(raw), 2)
    except ValueError:
        return None


def normalize_url(url: str, base_url: str = BASE_URL) -> str:
    if not url:
        return ""
    url = html.unescape(url).strip()
    if url.startswith("//"):
        url = "https:" + url
    elif not URL_RE.match(url):
        url = urljoin(base_url, url)
    return url


def normalize_image_url(url: str, base_url: str = BASE_URL) -> str:
    url = normalize_url(url, base_url)
    if not url:
        return ""
    # Prefer original jpg/png where possible, but keep valid webp URLs if only webp is present.
    url = url.replace("/webp-express/webp-images", "")
    if url.endswith(".webp"):
        url = url[:-5]
    url = THUMB_RE.sub("", url)
    return url


def should_skip_image(url: str) -> bool:
    low = url.lower()
    return any(part in low for part in SKIP_IMAGE_PARTS)


def unique_keep_order(items: Iterable[str]) -> list[str]:
    seen = set()
    out = []
    for item in items:
        item = clean_text(item)
        if not item or item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def slugify_sku_part(value: str, max_len: int = 28) -> str:
    """Make a compact suffix for duplicate SKU values."""
    value = clean_text(value)
    value = value.replace("№", "no")
    value = re.sub(r"[^0-9A-Za-zА-Яа-я._-]+", "-", value, flags=re.U).strip("-_.")
    return value[:max_len] or "variant"


def ensure_unique_seller_skus(rows: list[ProductRow]) -> list[ProductRow]:
    """Temu is stricter than WooCommerce: each row/SKU should be unique.

    Hygiene.bg sometimes uses the same SKU for several variations or even different
    products. Keep the first SKU as-is, and append a readable suffix to later duplicates.
    """
    seen: set[str] = set()
    used_counts: dict[str, int] = defaultdict(int)
    for row in rows:
        base = clean_text(row.seller_sku or row.listing_id or slug_from_url(row.product_url))
        if base and base not in seen:
            row.seller_sku = base
            seen.add(base)
            used_counts[base] += 1
            continue

        used_counts[base] += 1
        suffix_source = row.variation_variant_1 or row.item_name or str(used_counts[base])
        suffix = slugify_sku_part(suffix_source)
        candidate = f"{base}-{suffix}" if base else suffix
        i = 2
        while candidate in seen:
            candidate = f"{base}-{suffix}-{i}" if base else f"{suffix}-{i}"
            i += 1
        row.seller_sku = candidate
        seen.add(candidate)
    return rows


def get_meta(soup: BeautifulSoup, prop: str) -> str:
    tag = soup.find("meta", attrs={"property": prop}) or soup.find("meta", attrs={"name": prop})
    return clean_text(tag.get("content")) if tag and tag.get("content") else ""


def is_real_product_page(soup: BeautifulSoup) -> bool:
    """Return True only for actual WooCommerce product pages, not blog/category pages.

    Hygiene.bg uses root-level slugs for both products and blog posts, so URL shape alone
    is not enough. This guard prevents blog articles that contain product cards from being
    exported as products.
    """
    og_type = get_meta(soup, "og:type").lower()
    if og_type == "product":
        return True

    body_classes = set(soup.body.get("class", [])) if soup.body else set()
    if "single-product" in body_classes or "product-template-default" in body_classes:
        return True

    if soup.select_one("h1.product_title") and (
        soup.select_one("form.cart, form.variations_form, .product_meta, .woocommerce-product-gallery")
        or get_meta(soup, "product:price:amount")
    ):
        return True

    return False


def canonical_url(soup: BeautifulSoup, fallback: str = "") -> str:
    link = soup.find("link", rel="canonical")
    if link and link.get("href"):
        return normalize_url(link.get("href"))
    og_url = get_meta(soup, "og:url")
    return normalize_url(og_url or fallback)


def extract_title(soup: BeautifulSoup) -> str:
    # Use the WooCommerce product title only. Blog/article titles are intentionally ignored
    # by the product-page guard, but this also avoids mixing article titles into product rows.
    h1 = soup.select_one("h1.product_title")
    if h1:
        return clean_text(h1.get_text(" "), 500)
    return clean_text(get_meta(soup, "og:title") or (soup.title.get_text(" ") if soup.title else ""), 500)


def extract_brand(soup: BeautifulSoup) -> str:
    # Prefer Yoast/WooCommerce product metadata. Do NOT read .pwb-brands-in-loop because
    # that selector also appears in blog pages with product cards and caused product names
    # to be exported as brands.
    brand = get_meta(soup, "product:brand")
    if brand:
        return brand
    link = soup.select_one(".summary .pwb-single-product-brands a[title], .pwb-single-product-brands a[title]")
    if link and link.get("title"):
        return clean_text(link.get("title"))
    img = soup.select_one(".summary .pwb-single-product-brands img[alt], .pwb-single-product-brands img[alt]")
    return clean_text(img.get("alt")) if img and img.get("alt") else ""


def extract_sku(soup: BeautifulSoup) -> str:
    meta_id = get_meta(soup, "product:retailer_item_id")
    if meta_id:
        return meta_id
    for sku_el in soup.select(".sku_wrapper .sku, span.sku"):
        val = clean_text(sku_el.get_text(" "))
        if val and val.lower() not in {"няма", "n/a", "na"}:
            return val
    # REST shortlink contains product ID: ?p=1234
    shortlink = soup.find("link", rel="shortlink")
    if shortlink and shortlink.get("href"):
        m = re.search(r"[?&]p=(\d+)", shortlink.get("href"))
        if m:
            return m.group(1)
    return ""


def slug_from_url(url: str) -> str:
    path = urlparse(url).path.strip("/").split("/")[-1]
    return re.sub(r"[^A-Za-z0-9_-]+", "-", path).strip("-") or "product"


def extract_categories(soup: BeautifulSoup) -> list[str]:
    cats = []
    for a in soup.select(".posted_in a[rel='tag'], .posted_in a[href*='/cat/']"):
        txt = clean_text(a.get_text(" "))
        if txt and txt.lower() not in {"магазин", "shop"}:
            cats.append(txt)
    if not cats:
        for body_cls in soup.body.get("class", []) if soup.body else []:
            if body_cls.startswith("product_cat-"):
                cats.append(body_cls.replace("product_cat-", "").replace("-", " "))
    return unique_keep_order(cats)


def clean_multiline_text(element: Any) -> str:
    """Clean HTML into plain text while preserving paragraph/list boundaries."""
    if not element:
        return ""
    clone = BeautifulSoup(str(element), "lxml")
    for bad in clone.select("script, style, img, noscript, form, button"):
        bad.decompose()
    # Add line breaks around common block elements before extracting text.
    for br in clone.find_all("br"):
        br.replace_with("\n")
    for tag in clone.find_all(["p", "li", "tr", "h2", "h3", "h4"]):
        tag.insert_before("\n")
        tag.insert_after("\n")
    text = html.unescape(clone.get_text("\n"))
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    return "\n".join(unique_keep_order(lines))


def extract_description_and_notes(soup: BeautifulSoup) -> tuple[str, list[str]]:
    parts = []
    note_candidates = []

    # Short description first, then the full WooCommerce description tab. Keep full text;
    # do not truncate to 2000 chars because the user wants the full product description.
    for selector in (".woocommerce-product-details__short-description", "#tab-description", ".woocommerce-Tabs-panel--description"):
        element = soup.select_one(selector)
        if not element:
            continue
        text = clean_multiline_text(element)
        if text:
            parts.append(text)
            note_candidates.append(text)
        for li in element.select("li"):
            txt = clean_text(li.get_text(" "), 700)
            if txt and len(txt) > 3:
                note_candidates.append(txt)

    description = "\n\n".join(unique_keep_order(parts)).strip()
    # Fallback for products with no visible WooCommerce description tab.
    # Use meta description/OG description instead of leaving Temu description blank.
    if not description:
        fallback_description = clean_text(get_meta(soup, "og:description") or get_meta(soup, "description"))
        if fallback_description:
            description = fallback_description
            note_candidates.append(fallback_description)

    notes = []
    for candidate in note_candidates:
        # Split very long candidate into sentences for bullet points.
        splits = re.split(r"(?<=[.!?])\s+|\s*[•\-–]\s+", candidate.replace("\n", " "))
        for item in splits:
            item = clean_text(item, 700)
            if len(item) >= 10 and item not in notes:
                notes.append(item)
            if len(notes) >= 6:
                break
        if len(notes) >= 6:
            break
    return description, notes[:6]


def extract_gallery_images(soup: BeautifulSoup) -> list[str]:
    urls = []
    for a in soup.select(".woocommerce-product-gallery a[href]"):
        urls.append(normalize_image_url(a.get("href")))
    for img in soup.select(".woocommerce-product-gallery img"):
        for attr in ("data-large_image", "data-src", "src"):
            if img.get(attr):
                urls.append(normalize_image_url(img.get(attr)))
                break
    for meta in soup.find_all("meta", attrs={"property": "og:image"}):
        if meta.get("content"):
            urls.append(normalize_image_url(meta.get("content")))
    urls = [u for u in urls if u and not should_skip_image(u)]
    return unique_keep_order(urls)[:10]


def find_size(text: str) -> tuple[str, str, str]:
    """Return original size text, numeric content, normalized unit."""
    if not text:
        return "", "", ""
    m = SIZE_RE.search(text)
    if not m:
        return "", "", ""
    num = m.group("num").replace(",", ".")
    unit_raw = m.group("unit").lower().replace(".", "")
    if unit_raw.startswith("мл") or unit_raw.startswith("ml") or unit_raw.startswith("мил"):
        unit = "ml"
        display = f"{num} мл"
    elif unit_raw in {"л", "l"} or unit_raw.startswith("лит"):
        unit = "l"
        display = f"{num} л"
    elif unit_raw in {"кг", "kg"}:
        unit = "kg"
        display = f"{num} кг"
    else:
        unit = "g"
        display = f"{num} г"
    return display, num, unit


def classify_variant_type(label: str, value: str) -> str:
    combined = f"{label} {value}".lower()
    if any(x in combined for x in ["аромат", "мирис", "ухание", "fragrance", "scent"]):
        return "Аромат"
    if any(x in combined for x in ["цвят", "color", "colour"]):
        return "Цвят"
    if any(x in combined for x in ["туба", "обем", "разфас", "лит", "мл", "ml", "л.", "kg", "кг", "гр", "g."]):
        return "Обем"
    label = clean_text(label)
    return label or "Вариант"


def parse_variation_json(soup: BeautifulSoup) -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, str]]:
    form = soup.select_one("form.variations_form")
    if not form:
        return [], {}, {}
    raw = form.get("data-product_variations") or ""
    variations: list[dict[str, Any]] = []
    if raw:
        try:
            variations = json.loads(html.unescape(raw))
        except json.JSONDecodeError:
            logging.warning("Could not decode data-product_variations for %s", canonical_url(soup))
            variations = []

    option_labels: dict[str, dict[str, str]] = {}
    attr_labels: dict[str, str] = {}
    for select in form.select("select[name^='attribute_']"):
        attr = select.get("data-attribute_name") or select.get("name") or select.get("id") or ""
        label = ""
        tr = select.find_parent("tr")
        if tr:
            label_el = tr.select_one("th.label label, label")
            if label_el:
                label = clean_text(label_el.get_text(" "))
        if not label and select.get("id"):
            label = select.get("id").replace("pa_", "").replace("_", " ").title()
        attr_labels[attr] = label
        option_labels[attr] = {}
        for option in select.select("option[value]"):
            val = option.get("value", "")
            txt = clean_text(option.get_text(" "))
            if val:
                option_labels[attr][val] = txt
    return variations, option_labels, attr_labels


def extract_base_price(soup: BeautifulSoup) -> Optional[float]:
    meta_price = get_meta(soup, "product:price:amount")
    if meta_price:
        try:
            return round(float(meta_price.replace(",", ".")), 2)
        except ValueError:
            pass
    # Prefer current sale price if an <ins> price exists.
    ins = soup.select_one("p.price ins .woocommerce-Price-amount, .main-price ins .woocommerce-Price-amount")
    if ins:
        price = parse_price_text(ins.get_text(" "))
        if price is not None:
            return price
    price_el = soup.select_one("p.main-price, p.price, span.price")
    if price_el:
        return parse_price_text(str(price_el)) or parse_price_text(price_el.get_text(" "))
    return None


def extract_product_rows_from_html(page_html: str, source_url: str = "") -> list[ProductRow]:
    soup = BeautifulSoup(page_html, "lxml")
    if not is_real_product_page(soup):
        logging.info("Skipping non-product page: %s", source_url or canonical_url(soup))
        return []
    url = canonical_url(soup, source_url)
    title = extract_title(soup)
    base_sku = extract_sku(soup) or slug_from_url(url)
    brand = extract_brand(soup)
    categories = extract_categories(soup)
    category_text = " | ".join(categories)
    description, notes = extract_description_and_notes(soup)
    base_images = extract_gallery_images(soup)
    base_price = extract_base_price(soup)
    availability = get_meta(soup, "product:availability") or get_meta(soup, "og:availability")

    listing_id = base_sku or slug_from_url(url)
    keyword_attributes = " | ".join(unique_keep_order(categories + ([brand] if brand else [])))

    variations, option_labels, attr_labels = parse_variation_json(soup)
    rows: list[ProductRow] = []

    visible_variations = [v for v in variations if v.get("variation_is_visible", True) and v.get("variation_is_active", True)]
    if visible_variations:
        for idx, var in enumerate(visible_variations, start=1):
            attrs = var.get("attributes", {}) or {}
            var_pairs = []
            for attr_name, attr_slug in attrs.items():
                label = attr_labels.get(attr_name) or attr_name.replace("attribute_pa_", "").replace("attribute_", "").replace("_", " ").title()
                value_label = option_labels.get(attr_name, {}).get(attr_slug) or clean_text(str(attr_slug).replace("-", " "))
                var_type = classify_variant_type(label, value_label)
                var_pairs.append((var_type, value_label))

            if not var_pairs:
                var_pairs = [("Вариант", str(idx))]

            # For variable products, look at the actual variation value and product title.
            # Avoid scanning the full description to prevent unrelated capacities/weights from
            # being exported as net content.
            size_display, net_content, net_unit = find_size(" ".join(v for _, v in var_pairs) + " " + title)

            price = var.get("display_price")
            try:
                price = round(float(price), 2) if price is not None else base_price
            except Exception:
                price = base_price

            images = []
            image_info = var.get("image") or {}
            if image_info:
                for key in ("full_src", "url", "src"):
                    if image_info.get(key):
                        images.append(normalize_image_url(image_info[key]))
                        break
            images.extend(base_images)
            images = [u for u in unique_keep_order(images) if not should_skip_image(u)][:10]

            var_value = ", ".join(v for _, v in var_pairs if v)
            item_name = title
            if var_value and var_value.lower() not in title.lower():
                item_name = clean_text(f"{title} - {var_value}", 500)

            var_sku = clean_text(var.get("sku")) or f"{listing_id}-{idx}"
            row = ProductRow(
                product_url=url,
                listing_id=listing_id,
                seller_sku=var_sku,
                item_name=item_name,
                category=category_text,
                brand=brand,
                trademark=brand,
                variation_type_1=var_pairs[0][0],
                variation_variant_1=var_pairs[0][1],
                variation_type_2=var_pairs[1][0] if len(var_pairs) > 1 else "",
                variation_variant_2=var_pairs[1][1] if len(var_pairs) > 1 else "",
                item_notes=notes,
                item_description=description,
                keyword_attributes=keyword_attributes,
                images=images,
                map_price=price,
                net_content=net_content,
                total_net_content="",
                net_content_unit=net_unit,
                availability=availability,
                raw_variation=json.dumps(attrs, ensure_ascii=False),
            )
            rows.append(row)
        return rows

    # Simple product: infer product net content from the TITLE only.
    # Do not scan the full description here, because many non-liquid products mention
    # capacities/weights (e.g. vacuum bag capacity, machine weight) that are not actual
    # sellable variation values.
    size_display, net_content, net_unit = find_size(title)
    variation_type = "Обем" if size_display else "Вариант"
    variation_value = size_display or "Стандартен"

    row = ProductRow(
        product_url=url,
        listing_id=listing_id,
        seller_sku=base_sku,
        item_name=title,
        category=category_text,
        brand=brand,
        trademark=brand,
        variation_type_1=variation_type,
        variation_variant_1=variation_value,
        item_notes=notes,
        item_description=description,
        keyword_attributes=keyword_attributes,
        images=base_images,
        map_price=base_price,
        net_content=net_content,
        total_net_content="",
        net_content_unit=net_unit,
        availability=availability,
        raw_variation="",
    )
    return [row]


class Scraper:
    def __init__(self, base_url: str = BASE_URL, delay: float = 0.6, timeout: int = 30):
        self.base_url = base_url.rstrip("/") + "/"
        self.delay = delay
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})

    def get(self, url: str) -> str:
        time.sleep(self.delay)
        url = normalize_url(url, self.base_url)
        resp = self.session.get(url, timeout=self.timeout)
        resp.raise_for_status()
        resp.encoding = resp.apparent_encoding or "utf-8"
        return resp.text

    def discover_from_sitemap(self, sitemap_url: str) -> list[str]:
        logging.info("Reading sitemap: %s", sitemap_url)
        try:
            text = self.get(sitemap_url)
        except Exception as e:
            logging.warning("Failed sitemap %s: %s", sitemap_url, e)
            return []
        soup = BeautifulSoup(text, "xml")
        locs = [clean_text(loc.get_text()) for loc in soup.find_all("loc")]
        urls = []
        child_sitemaps = []
        for loc in locs:
            low = loc.lower()
            if low.endswith(".xml"):
                # Only product sitemaps should be traversed. Post/page/category sitemaps may
                # contain blog articles with product cards, which must not become product rows.
                if "product-sitemap" in low and "product_cat" not in low:
                    child_sitemaps.append(loc)
            elif self.is_probable_product_url(loc):
                urls.append(loc)
        for child in child_sitemaps:
            if child == sitemap_url:
                continue
            try:
                child_text = self.get(child)
                child_soup = BeautifulSoup(child_text, "xml")
                for loc in child_soup.find_all("loc"):
                    u = clean_text(loc.get_text())
                    if self.is_probable_product_url(u):
                        urls.append(u)
            except Exception as e:
                logging.warning("Failed child sitemap %s: %s", child, e)
        return unique_keep_order(urls)

    def discover_from_rest(self) -> list[str]:
        urls = []
        api_base = urljoin(self.base_url, "wp-json/wp/v2/product")
        for page in range(1, 200):
            params = {"per_page": 100, "page": page, "_fields": "link"}
            try:
                time.sleep(self.delay)
                resp = self.session.get(api_base, params=params, timeout=self.timeout)
                if resp.status_code in {400, 404}:
                    break
                resp.raise_for_status()
                items = resp.json()
                if not items:
                    break
                for item in items:
                    if item.get("link") and self.is_probable_product_url(item["link"]):
                        urls.append(item["link"])
                total_pages = int(resp.headers.get("X-WP-TotalPages", page))
                if page >= total_pages:
                    break
            except Exception as e:
                logging.warning("REST discovery stopped on page %s: %s", page, e)
                break
        return unique_keep_order(urls)

    def discover_from_category(self, category_url: str, max_pages: int = 300) -> list[str]:
        urls = []
        seen_pages = set()
        next_url = normalize_url(category_url, self.base_url)
        for _ in range(max_pages):
            if not next_url or next_url in seen_pages:
                break
            seen_pages.add(next_url)
            logging.info("Reading category page: %s", next_url)
            try:
                page_html = self.get(next_url)
            except Exception as e:
                logging.warning("Failed category page %s: %s", next_url, e)
                break
            soup = BeautifulSoup(page_html, "lxml")
            for a in soup.select("li.product a.woocommerce-LoopProduct-link, li.product a[href]"):
                href = a.get("href")
                if href:
                    u = normalize_url(href, self.base_url)
                    if self.is_probable_product_url(u):
                        urls.append(u)
            next_link = soup.find("link", rel="next") or soup.select_one("a.next.page-numbers")
            next_url = normalize_url(next_link.get("href"), self.base_url) if next_link and next_link.get("href") else ""
        return unique_keep_order(urls)

    def discover_from_home_categories(self) -> list[str]:
        try:
            soup = BeautifulSoup(self.get(self.base_url), "lxml")
        except Exception as e:
            logging.warning("Home discovery failed: %s", e)
            return []
        cat_urls = []
        for a in soup.select("a[href*='/cat/']"):
            href = normalize_url(a.get("href"), self.base_url)
            if "/cat/" in href:
                cat_urls.append(href)
        product_urls = []
        for cat in unique_keep_order(cat_urls):
            product_urls.extend(self.discover_from_category(cat, max_pages=100))
        return unique_keep_order(product_urls)

    def discover_all_site(self) -> list[str]:
        candidates = [
            urljoin(self.base_url, "sitemap_index.xml"),
            urljoin(self.base_url, "product-sitemap.xml"),
        ] + [urljoin(self.base_url, f"product-sitemap{i}.xml") for i in range(1, 31)]
        urls = []
        for sitemap in candidates:
            urls.extend(self.discover_from_sitemap(sitemap))
        urls = unique_keep_order(urls)
        if urls:
            return urls
        logging.warning("No products found from sitemap. Trying WordPress REST API.")
        urls = self.discover_from_rest()
        if urls:
            return urls
        logging.warning("No products found from REST API. Trying category crawl.")
        return self.discover_from_home_categories()

    def is_probable_product_url(self, url: str) -> bool:
        url = normalize_url(url, self.base_url)
        parsed = urlparse(url)
        if not parsed.netloc.endswith("hygiene.bg"):
            return False
        path = parsed.path.strip("/")
        if not path or "/" in path.rstrip("/"):
            # Hygiene product URLs are usually root-level slugs; exclude known taxonomy/pages.
            if not path.startswith("product/"):
                return False
        bad_fragments = [
            "cat/", "brand/", "label/", "cart", "checkout", "my-profile", "wp-content", "wp-json",
            "promocii", "kontakti", "contacts", "shop/page", "page/", "feed", "sitemap",
        ]
        if any(bad in path.lower() for bad in bad_fragments):
            return False
        return True

    def scrape_products(self, urls: list[str], limit: int = 0) -> list[ProductRow]:
        rows: list[ProductRow] = []
        urls = unique_keep_order(urls)
        if limit and limit > 0:
            urls = urls[:limit]
        for i, url in enumerate(urls, start=1):
            logging.info("[%s/%s] Scraping %s", i, len(urls), url)
            try:
                html_text = self.get(url)
                product_rows = extract_product_rows_from_html(html_text, url)
                product_rows = [r for r in product_rows if r.item_name and r.images and r.map_price is not None]
                rows.extend(product_rows)
            except Exception as e:
                logging.exception("Failed product %s: %s", url, e)
        return rows


def rows_from_local_html(paths: list[str], limit: int = 0) -> list[ProductRow]:
    rows: list[ProductRow] = []
    for path in paths[: limit or None]:
        text = Path(path).read_text(encoding="utf-8", errors="ignore")
        rows.extend(extract_product_rows_from_html(text, Path(path).as_uri()))
    return rows


def build_column_map(ws) -> dict[str, list[int]]:
    cols: dict[str, list[int]] = defaultdict(list)
    for cell in ws[3]:
        if cell.value:
            cols[str(cell.value)].append(cell.column)
    return cols


def write_temu_template(template_path: str, output_path: str, rows: list[ProductRow]) -> None:
    shutil.copyfile(template_path, output_path)
    wb = load_workbook(output_path)
    if "Template" not in wb.sheetnames:
        raise RuntimeError("Template sheet not found in workbook")
    ws = wb["Template"]

    # Clear old user data under the header area, preserving the first 4 template rows.
    if ws.max_row > 4:
        ws.delete_rows(5, ws.max_row - 4)

    col_map = build_column_map(ws)

    def set_first(row_idx: int, field_name: str, value: Any) -> None:
        if field_name in col_map and col_map[field_name]:
            ws.cell(row_idx, col_map[field_name][0]).value = value

    def set_occurrences(row_idx: int, field_name: str, values: Iterable[Any]) -> None:
        for col, value in zip(col_map.get(field_name, []), values):
            ws.cell(row_idx, col).value = value

    for row_idx, item in enumerate(rows, start=5):
        set_first(row_idx, "listing_id", item.listing_id)
        set_first(row_idx, "seller_sku", item.seller_sku)
        set_first(row_idx, "item_name", item.item_name)
        set_first(row_idx, "category", item.category)
        set_first(row_idx, "reference_link", item.product_url)
        set_first(row_idx, "brand", item.brand)
        set_first(row_idx, "trademark", item.trademark)
        set_first(row_idx, "Variation_type_1", item.variation_type_1)
        set_first(row_idx, "Variation_Variants_1", item.variation_variant_1)
        set_first(row_idx, "Variation_type_2", item.variation_type_2)
        set_first(row_idx, "Variation_Variants_2", item.variation_variant_2)
        set_occurrences(row_idx, "item_note", item.item_notes)
        set_first(row_idx, "item_description", item.item_description)
        set_first(row_idx, "Keyword_attributes", item.keyword_attributes)
        if item.images:
            set_first(row_idx, "main_image_url", item.images[0])
            set_occurrences(row_idx, "other_image_url", item.images[1:10])
        set_first(row_idx, "map_price", item.map_price)
        # Per user request: quantity, package dimensions, weight and shipping template remain blank.
        set_first(row_idx, "Net_content", item.net_content)
        set_first(row_idx, "Total_net_content", item.total_net_content)
        set_first(row_idx, "Net_content_unit", item.net_content_unit)
        # Keep item_length/item_width/item_height/item_weight/item_volume_unit/item_weight_unit blank.

    wb.save(output_path)


def write_raw_csv(path: str, rows: list[ProductRow]) -> None:
    fieldnames = list(asdict(rows[0]).keys()) if rows else [f.name for f in ProductRow.__dataclass_fields__.values()]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            data = asdict(row)
            data["item_notes"] = " | ".join(row.item_notes)
            data["images"] = " | ".join(row.images)
            writer.writerow(data)


def main() -> int:
    parser = argparse.ArgumentParser(description="Scrape hygiene.bg products into TEMU upload template")
    parser.add_argument("--mode", choices=["all_site", "single_product", "category", "sitemap"], default="all_site")
    parser.add_argument("--url", default="https://hygiene.bg/sitemap_index.xml", help="URL for single_product/category/sitemap/all_site seed")
    parser.add_argument("--base-url", default=BASE_URL)
    parser.add_argument("--template", default="TEMU_GENERAL_UPLOAD_TEMPLATE.xlsx")
    parser.add_argument("--output", default="TEMU_HYGIENE_UPLOAD.xlsx")
    parser.add_argument("--raw-output", default="hygiene_raw_export.csv")
    parser.add_argument("--limit", type=int, default=0, help="0 = no limit")
    parser.add_argument("--delay", type=float, default=0.6)
    parser.add_argument("--local-html", nargs="*", help="Optional local HTML files for offline testing")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args()

    logging.basicConfig(level=getattr(logging, args.log_level.upper(), logging.INFO), format="%(levelname)s: %(message)s")

    if args.local_html:
        rows = rows_from_local_html(args.local_html, limit=args.limit)
    else:
        scraper = Scraper(base_url=args.base_url, delay=args.delay)
        if args.mode == "single_product":
            urls = [args.url]
        elif args.mode == "category":
            urls = scraper.discover_from_category(args.url)
        elif args.mode == "sitemap":
            urls = scraper.discover_from_sitemap(args.url)
        else:
            urls = scraper.discover_all_site()
        logging.info("Discovered %s product URLs", len(urls))
        rows = scraper.scrape_products(urls, limit=args.limit)

    rows = ensure_unique_seller_skus(rows)

    logging.info("Exporting %s SKU rows", len(rows))
    if not rows:
        logging.error("No product rows were extracted. Check the URL/mode or website availability.")
        return 2
    write_temu_template(args.template, args.output, rows)
    write_raw_csv(args.raw_output, rows)
    logging.info("Saved %s and %s", args.output, args.raw_output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
